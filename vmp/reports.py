"""Review and export helpers for analyzed media plans."""

from __future__ import annotations

from pathlib import Path

from .core.i18n import tr
from .metadata import has_exif_datetime_values, missing_exif_suggestion
from .core.models import ActionKind, AnalysisResult, AppSettings, MediaKind, MediaPlan, PlanStatus
from .planner import crf_for_video, video_bucket_label


def plan_action_summary(plan: MediaPlan) -> str:
    """Return a concise human-readable summary of planned work."""
    if not plan.actions:
        return tr("Scan ausstehend")
    result = plan.analysis
    fallback_note = ""
    if "System:FileName" in result.resolved.source and not has_exif_datetime_values(result.metadata):
        fallback_note = tr("Datum aus Dateiname geparst")
    if result.item.kind == MediaKind.IMAGE and result.item.suffix in {".jpg", ".jpeg"}:
        action_kinds = {action.kind for action in plan.actions}
        if ActionKind.IMAGE_CONVERT not in action_kinds and ActionKind.VIDEO_TRANSCODE not in action_kinds:
            parts: list[str] = []
            if result.status == PlanStatus.DONE:
                parts.append(tr("Bereits verarbeitet"))
            parts.append(tr("Nur JPEG-Metadaten"))
            for action in plan.actions:
                if action.kind == ActionKind.RENAME:
                    parts.append(action.description)
            if fallback_note:
                parts.append(fallback_note)
            return "; ".join(parts)
    parts = [action.description for action in plan.actions]
    if fallback_note:
        parts.append(fallback_note)
    return "; ".join(parts)


def preview_row(plan: MediaPlan, settings: AppSettings) -> list[str]:
    """Build one dry-run preview row."""
    result = plan.analysis
    resolved = result.resolved
    actions = plan_action_summary(plan)
    bucket = ""
    if result.item.kind == MediaKind.VIDEO:
        bucket = f"{video_bucket_label(result, settings)} / CRF {crf_for_video(result, settings)}"
    return [
        result.status.value.upper(),
        result.item.kind.value,
        resolved.confidence.value,
        str(result.item.relative_path),
        resolved.local_dt.isoformat(sep=" ") if resolved.local_dt else "",
        str(resolved.offset) if resolved.offset else "",
        resolved.utc_dt.isoformat(sep=" ") if resolved.utc_dt else "",
        bucket,
        str(plan.final_path.name) if plan.final_path else "",
        actions,
        "\n".join(result.warnings),
    ]


def debug_row(plan: MediaPlan) -> list[str]:
    """Build one core datetime debug row."""
    result = plan.analysis
    resolved = result.resolved
    return [
        str(result.item.relative_path),
        result.status.value.upper(),
        resolved.local_dt.isoformat(sep=" ") if resolved.local_dt else "",
        ", ".join(resolved.local_sources),
        str(resolved.offset) if resolved.offset else "",
        ", ".join(resolved.offset_sources),
        resolved.utc_dt.isoformat(sep=" ") if resolved.utc_dt else "",
        ", ".join(resolved.utc_sources),
        "\n".join(resolved.notes),
    ]


def missing_exif_rows(results: list[AnalysisResult]) -> list[list[str]]:
    """Return rows for files without core EXIF capture dates."""
    rows: list[list[str]] = []
    for result in results:
        if result.item.kind == MediaKind.VIDEO and (
            result.resolved.local_dt is not None
            or result.resolved.utc_dt is not None
            or result.resolved.local_date_only
        ):
            continue
        if has_exif_datetime_values(result.metadata):
            continue
        rows.append(
            [
                str(result.item.relative_path),
                result.item.kind.value,
                result.resolved.local_dt.isoformat(sep=" ") if result.resolved.local_dt else "",
                result.resolved.confidence.value,
                missing_exif_suggestion(result),
            ]
        )
    return rows


def vacation_span_warning(results: list[AnalysisResult], max_weeks: int) -> str | None:
    """Return a warning message when resolved dates span more than max_weeks."""
    if max_weeks <= 0:
        return None
    dates = [
        result.resolved.local_dt
        for result in results
        if result.resolved.local_dt is not None and not result.resolved.local_date_only
    ]
    if len(dates) < 2:
        return None
    first = min(dates)
    last = max(dates)
    span_days = (last - first).days
    if span_days <= max_weeks * 7:
        return None
    return tr("Die erkannten Aufnahmedaten spannen {days} Tage ({first} bis {last}).").format(
        days=span_days, first=f"{first:%Y-%m-%d}", last=f"{last:%Y-%m-%d}"
    )


def export_excel_report(plans: list[MediaPlan], output_path: Path, settings: AppSettings) -> None:
    """Export dry-run, core debug, and missing-EXIF review sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    preview_sheet = workbook.active
    preview_sheet.title = "DryRunPreview"
    debug_sheet = workbook.create_sheet("CoreDatetimeDebug")
    missing_sheet = workbook.create_sheet("MissingExifReview")

    def sanitize_cell(value: str) -> str:
        """Prevent accidental Excel formula interpretation of cell text."""
        return f"'{value}" if isinstance(value, str) and value.startswith(("=", "+", "@", "-")) else value

    def write_sheet(sheet: object, headers: list[str], rows: list[list[str]]) -> None:
        """Write a styled worksheet."""
        assert hasattr(sheet, "append")
        sheet.append(headers)  # type: ignore[attr-defined]
        for row in rows:
            sheet.append([sanitize_cell(value) for value in row])  # type: ignore[attr-defined]
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:  # type: ignore[index]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        max_col = len(headers)
        max_row = sheet.max_row  # type: ignore[attr-defined]
        sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"  # type: ignore[attr-defined]
        sheet.freeze_panes = "A2"  # type: ignore[attr-defined]
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1])
            for row_idx in range(2, max_row + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value  # type: ignore[attr-defined]
                max_len = max(max_len, len("" if value is None else str(value)))
            sheet.column_dimensions[letter].width = min(max(12, max_len + 2), 80)  # type: ignore[attr-defined]

    write_sheet(
        preview_sheet,
        ["Status", "Type", "Confidence", "File", "Local", "Offset", "UTC", "Bucket", "Rename", "Actions", "Warnings"],
        [preview_row(plan, settings) for plan in plans],
    )
    write_sheet(
        debug_sheet,
        ["File", "Status", "LOCAL", "LOCAL Source", "OFFSET", "OFFSET Source", "UTC", "UTC Source", "Notes"],
        [debug_row(plan) for plan in plans],
    )
    write_sheet(
        missing_sheet,
        ["File", "Type", "Resolved Local", "Confidence", "Suggestion"],
        missing_exif_rows([plan.analysis for plan in plans]),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

def resolution_class(width: int | None, height: int | None) -> str:
    """Return a coarse, settings-independent resolution label for display.

    Classifies by the long edge into 4K/QHD/FHD/HD/SD. Unlike ``video_bucket``
    (which drives encoding and uses the configured thresholds), this is a fixed
    scale meant purely for the table's Codec column. Returns "" when unknown.
    """
    long_edge = max(width or 0, height or 0)
    if long_edge <= 0:
        return ""
    if long_edge >= 3200:
        return "4K"
    if long_edge >= 2200:
        return "QHD"
    if long_edge >= 1600:
        return "FHD"
    if long_edge >= 1120:
        return "HD"
    return "SD"
def display_video_codec(codec: str | None) -> str:
    """Normalize a raw codec / compressor id to a friendly display name.

    Maps the various HEVC/AVC spellings (``hevc``, ``hvc1``, ``h265``, …) to
    ``x265``/``x264`` so the table always shows the codec rather than a
    container id like ``MOV``. Unknown values are returned trimmed as-is.
    """
    if not codec:
        return ""
    key = codec.strip().lower()
    if any(token in key for token in ("hevc", "h265", "h.265", "x265", "hvc1", "hev1")):
        return "x265"
    if any(token in key for token in ("avc", "h264", "h.264", "x264")):
        return "x264"
    if key in ("av1", "av01") or "av01" in key:
        return "AV1"
    if "vp9" in key:
        return "VP9"
    if "vp8" in key:
        return "VP8"
    if "mpeg4" in key or key in ("mp4v", "xvid", "divx"):
        return "MPEG-4"
    if "mpeg2" in key:
        return "MPEG-2"
    return codec.strip()
